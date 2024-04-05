from openpyxl import Workbook, load_workbook
from Diploma_1 import Person
from datetime import datetime


def load_data(filename):
    data = []
    wb = load_workbook(filename)
    sh = wb.active
    for row in sh.iter_rows(values_only=True):
        p = Person(*row)
        data.append(p)
        born = lambda gend: ('Народилася' if gend == 'ж' else 'Народився')
        died = lambda gend: ('Померла' if gend == 'ж' else 'Помер')
        years = lambda age: 'роки' if (
                    str(age)[-1] in '234' and (True if len(str(age)) == 1 else str(age)[-2] != '1')) else 'рік' if \
        str(age)[-1] == '1' else 'років'

        print(
            f"{p.first_name} {p.middle_name or ''} {p.last_name or ''}, {p.calculate_age()} {years(p.calculate_age())}, {p.gender}. {born(p.gender)}: {p.birth_date}. {died(p.gender)}: {p.death_date or 'N/A'}"
             )
    return data


def save_data(filename, data):
    wb = Workbook()
    sh = wb.active
    for person in data:
        sh.append([person.first_name, person.last_name, person.middle_name, person.birth_date, person.death_date, person.gender])
    wb.save(filename)


def search(data, word):
    results = []
    for person in data:
        if word.lower() in person.first_name.lower() or word.lower() in (
                person.last_name or '').lower() or word.lower() in (person.middle_name or '').lower():
            results.append(person)
    return results


if __name__ == '__main__':
    people = []
    while True:
        print("1. Завантажити дані")
        print("2. Зберегти дані")
        print("3. Додати новий запис")
        print("4. Пошук")
        print("5. Вихід")
        answer = input("Введіть номер пункту: ")

        if answer == '1':
            try:
                filename = input("Введіть назву файлу: ")
                people = load_data(filename)
            except Exception:
                print('***Помилка вводу!***')

        elif answer == '2':
            try:
                filename = input("Введіть бажану назву файлу: ")
                save_data(filename, people)
            except Exception:
                print('***Помилка вводу!***')
        elif answer == '3':
            try:
                first_name = input("Введіть ім'я: ")
                if first_name.isdigit():
                    raise ValueError('***Помилка вводу!***')

                last_name = input("Введіть прізвище (за бажанням): ")
                if last_name.isdigit():
                    raise ValueError('***Помилка вводу!***')

                middle_name = input("Введіть по-батькові (за бажанням): ")
                if middle_name.isdigit():
                    raise ValueError('***Помилка вводу!***')

                birth_date = input("Введіть дату народження: ")
                birth_date = birth_date.replace('.', '/').replace('-', '/').replace(' ', '/')
                death_date = input("Введіть дату смерті (за наявності): ")
                if death_date:
                    death_date = death_date.replace('.', '/').replace('-', '/').replace(' ', '/')

                gender = input("Введіть стать: ")

                if gender.lower() not in ('м', 'ж'):
                    raise ValueError('***Помилка вводу!***')


                p = Person(first_name, last_name, middle_name, birth_date, death_date, gender)
                people.append(p)
            except ValueError as e:
                print(e)
        elif answer == '4':
            query = input("Enter your search query: ")
            results = search(people, query)
            born = lambda gend:('Народилася' if gend == 'ж' else 'Народився')
            died = lambda gend:('Померла' if gend == 'ж' else 'Помер')
            years = lambda age: 'роки' if (str(age)[-1] in '234' and (True if len(str(age)) == 1 else str(age)[-2] != '1')) else 'рік' if str(age)[-1] == '1' else 'років'

            for person in results:
                print(
                    f"{person.first_name} {person.middle_name or ''} {person.last_name or ''}  {person.calculate_age()} {years(person.calculate_age())}, {person.gender}. {born(person.gender)}: {person.birth_date}. {died(person.gender)}: {person.death_date or 'N/A'}")
        elif answer == '5':
            break