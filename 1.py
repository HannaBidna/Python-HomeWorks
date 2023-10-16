
from datetime import date

import openpyxl
import Diploma_1
from Diploma_1 import Person


person_1 = Person('Ліна', '16-03-1930', '', 'ж', 'Костенко', '')
person_2 = Person('Василь', '06-01-1938', '04/09/1985', 'ч', 'Стус', 'Семенович')


def load_data(filename):
    data = []
    wb = openpyxl.load_workbook(filename)
    sh = wb.active
    for row in sh.iter_rows(values_only=True):
        p = Person(*row)
        data.append(p)
    return data


def open_file():
    wb = openpyxl.load_workbook('test.xlsx')
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    data = []
    for row in range(2, rows + 1):
        string = ''
        for col in range(1, cols + 1):
            cell = sheet.cell(row=row, column=col)
            string += str(cell.value) + ' '
        data.append(string.strip())
        return data
    print(data)
    wb.save('test.xlsx')
    wb.close


def create_file(filename, data):
    wb = openpyxl.workbook()
    sh = wb.active
    for person in data:
        sh.append([person.name, person.birth_date, person.death_date, person.gender, person.surname, person.second_name])
    wb.save(filename)


def search(data, word):
    results = []
    for person in data:
        if word.lower() in person.name.lower() or word.lower() in (person.surname or '').lower() or word.lower() in (person.second_name or '').lower():
            results.append(person)
            return results
        else:
            print('Не знайдено!')


def main():
    if __name__ == '__main__':
        while True:
            people = []
            print('Виберіть необхідний пункт меню:', '1 - Завантажити дані ', '2 - Створити новий файл',
              '3 - Додати запис', '4 - Знайти дані', '5 - Вихід', sep='\n')
            answer = input()
            if answer == '1':
                filename = str(input("Введіть назву файлу: "))
                people = load_data(filename)
            elif answer == '2':
                filename = str(input("Введіть назву файлу: "))
                create_file(filename, people)
            elif answer == '3':
                name = input("Введіть імя: ")
                birth_date = input("Введіть дату народження: ")
                death_date = input("Введіть дату смерті (за наявності): ")
                gender = input("Введіть стать: ")
                surname = input("Введіть прізвище (за бажанням): ")
                second_name = input("Введіть по-батькові (за бажанням): ")
                p = Person(name, birth_date, death_date, gender, surname,  second_name)
                people.append(p)
            elif answer == '4':
                word = input("Введіть слово для пошуку: ")
                results = search(people, word)
                for person in results:
                    print(
                        f"{person.name} {person.second_name or ''} {person.surname or ''} {person.calculate_age()}років, {person.gender}. Народився: {person.birth_date}. Помер: {person.death_date or '--'}")
            elif answer == '5':
                break
            else:
                print('Помилка!')


main()
